"""
NETT = Narrative Entity Type(s) to Tables

This program takes as input:
* a corpus of relational tables, either in CSV (or Excel)
  or in a specific JSON format
  (the files are allowed to be inside one or multiple .tar archives);
  when no corpus is specified a small default corpus is being used
  => note that all tables smaller than 3x3 are rigorously filtered out
* optionally: a list of one or more entity types occuring in the narrative that
    is to be grounded (ideally as Wikidata ID's or DBpedia class names,
    otherwise as literal strings then being mapped to Wikidata ID's)
    => you may also supply ids/names of entities instead of entity types,
       in that case they are simply resolved to the entity type they're an
       instance of and the output tables restricted to those containing a
       string equal or similar to that entity's name # ToDo: this feature!
* various settings as parameters, see --help, also for more details
  on the 4 different modes of NETT!
"""

from __future__ import annotations
from typing import List, Dict, Any, Iterator
import json  # https://docs.python.org/3/library/json.html
import argparse
import re
import os
import sys
import csv  # https://docs.python.org/3/library/csv.html
import tarfile  # https://docs.python.org/3/library/tarfile.html
import tempfile  # https://docs.python.org/3/library/tempfile.html

from numpy import transpose

from WikidataItem import WikidataItem

# Import the fundamental approaches
#     #1 "Using Textual Surroundings"
#     #2 "Using Attribute Names"
#     #3 "Using Attribute Extensions"
# from the respective Python files:
from filter_nouns_with_heuristics import filter_nouns_with_heuristics_as_dict
from attr_names_to_ontology_class import attr_names_to_ontology_class
from attr_extension_to_ontology_class import attr_extension_to_ontology_class
from attr_extension_to_ontology_class_web_search import\
	attr_extension_to_ontology_class_web_search_list_onto_as_dict

from nett_map_dbpedia_classes_to_wikidata import\
	get_dbpedia_classes_mapped_to_wikidata
from nett_map_dbpedia_properties_to_sbert_vectors import\
	initalize_dbpedia_properties_mapped_to_SBERT_vector,\
	get_dbpedia_properties_mapped_to_SBERT_vector

DEBUG = True

def normalize(dct: Dict[WikidataItem, float])\
	-> Dict[WikidataItem, float]:
	if dct is not None:
		min_value: float = min(dct.values())
		max_value: float = max(dct.values())
		return {w: (f - min_value) / (max_value - min_value)\
				for w, f in dct.items()}
	else:
		return None


def combine3(dct1: Dict[WikidataItem, float], weight1: float,\
			dct2: Dict[WikidataItem, float], weight2: float,\
			dct3: Dict[WikidataItem, float], weight3: float)\
			-> List[Tuple[float, WikidataItem]]:
	allWikidataItems = set(dct1.keys())\
						.union(set(dct2.keys()))\
						.union(set(dct3.keys()))
	result: List[Tuple[float, WikidataItem]] =\
		[(weight1 * dct1.get(wi, 0.0) +\
		 weight2 * dct2.get(wi, 0.0) +\
		 weight3 * dct3.get(wi, 0.0),\
		 wi) for wi in allWikidataItems]
	result.sort(key=lambda tuple: tuple[0], reverse=True)
	return result


def debug_dict_sorted(d: Dict[WikidataItem, float]) -> str:
	"""
	A debug-printable short string representation of a
	Dict[WikidataItem, float] dictionary, where the floats are to be
	interpreted as match scores (bigger score = better match).
	"""
	if len(d) <= 2:  # empy dictionary or dictionary with <= 2 key value pairs:
		return str(d)
	else:
		max_key: WikidataItem = None
		max_value: float = float('-inf')
		for key in d:
			value: float = d[key]
			if value > max_value:
				max_key = key
				max_value = value
		return "{" + str(max_key) + ": " + str(max_value) + ", ...}"


class Table:
	def __init__(self, surroundingText: str, headerRow: List[str],\
				 columns: List[List[str]]):
		"""
		Initalize a new Table that has already been parsed.

		If the `columns` haven't been parsed for a `headerRow` yet,
		set headerRow=[] and
		call the parse_header_row() method after construction!

		`headerRow` may be `[]` but never `None`!
		"""

		self.surroundingText = surroundingText  # (1) Using Textual Surroundings
		self.headerRow = headerRow  # (2) Using Attribute Names
		self.columns = columns  # (3) Using Attribute Extensions

	def width(self) -> int:
		"""
		Returns the width of this Table, i.e. the number of columns.
		"""
		return len(self.columns)  # width == # of columns

	def min_height(self, includingHeaderRow=True) -> int:
		"""
		Returns the height of this Table, i.e. the number of rows.
		Including the header row by default (if this Table has one, that is).
		Set includingHeaderRow=False to exclude a possible header row.

		The value returned is different to that of max_height() only if
		the columns don't all have the same height, which they should have.
		"""
		return min(len(col) for col in self.columns) +\
			(1 if includingHeaderRow and self.headerRow != [] else 0)

	def max_height(self, includingHeaderRow=True) -> int:
		"""
		Returns the height of this Table, i.e. the number of rows.
		Including the header row by default (if this Table has one, that is).
		Set includingHeaderRow=False to exclude a possible header row.

		The value returned is different to that of min_height() only if
		the columns don't all have the same height, which they should have.
		"""
		return max(len(col) for col in self.columns) +\
			(1 if includingHeaderRow and self.headerRow != [] else 0)

	def pretty_print(self, maxNumberOfTuples=6, maxColWidth=25,\
		maxTotalWidth=180) -> str:
		"""
		A pretty printable version of this Table, e.g.:

		Restaurant Name       | Rating | Price | Reviews
		------------------------------------------------
		Aquarius              |        | $$    | 0
		BIG & little's        |        | $     | 0
		Brown Bag Seafood Co. |        | $$    | 0
		...                   | ...    | ...   | ...

		(in this example, maxNumberOfTuples=3)
		"""

        # The width (number of characters) of each column, only regarding
        #   the first `maxNumberOfTuples` rows each:
		columnWidths: List[int] = list(map(\
			lambda column: max(map(\
				lambda cell: len(cell),\
				column[0:maxNumberOfTuples]\
			)),\
			self.columns\
			))
		# If a header name is longer than corresponding column width, it
		#   has to be increased further:
		if self.headerRow != []:
			columnWidths = list(\
				map(\
					lambda tuple: max(tuple[0], tuple[1]),\
					zip(columnWidths, map(lambda header: len(header),\
						self.headerRow))\
					)\
				)
		# Every column shall have a width of at least 3:
		columnWidths = [max(3, cw) for cw in columnWidths]
		# Every column shall have a width of at most `maxColWidth`:
		columnWidths = [min(maxColWidth, cw) for cw in columnWidths]

		def print_row(cells: List[str], widths: List[int]) -> str:
			res: str = ""
			for i in range(0, len(cells)):
				width = widths[i]
				res += cells[i][:width] + " " * (width-len(cells[i]))
				if i != len(cells)-1: res += " | "
			res += "\n"
			return res

		result: str = ""

		if self.headerRow != []:
			result += print_row(self.headerRow, columnWidths)
		else:  # no header row to print => print empty header row:
			result += print_row([""] * len(columnWidths), columnWidths)

		totalWidth = sum(columnWidths) + 3*(len(columnWidths)-1)
		result += "-" * totalWidth + "\n" # "-----------------------------"

		for y in range(0, min(maxNumberOfTuples, len(self.columns[0]))):
			result += print_row(\
				[self.columns[x][y] for x in range(0, len(self.columns))],\
				 columnWidths)

		if maxNumberOfTuples < len(self.columns[0]):  # print "... | ..." row:
			result += print_row(["..."] * len(columnWidths), columnWidths)

		if maxTotalWidth > 0:
			result = "\n".join(map(lambda line: line\
				if len(line) <= maxTotalWidth\
				else line[:maxTotalWidth-4] + " ...", result.splitlines()))

		return result


	# (possible ToDo: quoting?!)
	def as_csv(self, with_header=True, max_number_of_rows: int = -1) -> str:
		self_min_height: int = self.min_height(includingHeaderRow=False)

		if max_number_of_rows == -1:
			max_number_of_rows = self_min_height

		self_as_csv: str = ""

		if with_header and self.headerRow != []:
			self_as_csv += ",".join(\
				self.headerRow[col_index]\
				for col_index in range(0, len(self.headerRow))) + "\n"

		for row_index in range(0, min(max_number_of_rows, self_min_height)):
			self_as_csv += ",".join(\
				self.columns[col_index][row_index]\
				for col_index in range(0, self.width())) + "\n"

		return self_as_csv

	def parse_header_row(self) -> bool:
		"""
		When self.headerRow == []:
		    Tries to to identify a header row in self.columns using the
		    heuristic used by the csv.Sniffer().has_header() method
		    (cf. https://docs.python.org/3/library/csv.html).
		    If found, the header row is deleted from self.columns and
		    self.headerRow is set instead.
		    Returns True when a header row was found and False otherwise.
		When self.headerRow != []:
		    This method does nothing and returns True.

		Important note:
		When using one of the parseXXX() class methods,
		calling this function is NOT necessary!!
		"""
		if self.headerRow == []:
			self_as_csv_sample: str =\
				self.as_csv(with_header=False, max_number_of_rows=20)
			# From https://docs.python.org/3/library/csv.html,
			# regarding csv.Sniffer.has_header(sample):
			# "Twenty rows after the first row are sampled; if more than half
			#  of columns + rows meet the criteria, True is returned."
			if csv.Sniffer().has_header(self_as_csv_sample):
				# Set the header row to the first row:
				self.headerRow = [self.columns[col_index][0] for\
					col_index in range(0, len(self.columns))]
				# Delete the header row from self.columns where it
				#   does not belong:
				for col_index in range(0, len(self.columns)):
					del self.columns[col_index][0]
				return True
			else:
				return False  # csv.Sniffer().has_header() found no header.
		else:
			return True  # There already is a header row (parsed).


	# Correctly regex-matching URLs and email addresses is actually much
	# harder than one might think, read:
	#
	# Email addresses:
	# * https://stackoverflow.com/questions/9238640/
	#     how-long-can-a-tld-possibly-be
	# * https://stackoverflow.com/questions/201323/
	#     how-can-i-validate-an-email-address-using-a-regular-expression
	# * http://www.ex-parrot.com/~pdw/Mail-RFC822-Address.html
	# * https://en.wikipedia.org/wiki/Email_address#Local-part
	# * http://emailregex.com/
	#
	# URLs:
	# * https://stackoverflow.com/questions/3809401/
	#     what-is-a-good-regular-expression-to-match-a-url
	# * https://mathiasbynens.be/demo/url-regex:

	PHONE_NUMBER_REGEX = r"\+?\d[\d -]+\d"
	URL_REGEX = r"^(https:|http:|www\.)\S*"  # https://regex101.com/r/S2CbwM/1
	EMAIL_REGEX = r".+@.+"
	#EMAIL_REGEX = r".+@.+\..+"
	# -> problem: "jsmith@[IPv6:2001:db8::1]" is a valid email without any dots!
	#EMAIL_REGEX = r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)"
	# -> source: emailregex.com
	# -> problem: !#$%&'*+-/=?^_`{|}~ are allowed characters in emails too!
	NUMERIC_REGEX = r"[\+\-]?[\d., _]+"  # r"[\d.,]+"
	GEO_REGEX = r"""[\+\-\d.,;'" NESW]+"""
	LONG_REGEX = r".{50,}"
	TIME_REGEX = r"\d\d:\d\d(:\d\d)?"  # e.g. '00:15' or '00:15:00'
	DATE_REGEX = r"(\d\d? \w{3,} \d\d\d\d)|(\w{3,} \d\d?, \d\d\d\d)"
	# e.g. '25 Jan 1893' or 'August 31, 2006'
	EMPTY_REGEX = r"\s*"  # e.g. '' or ' ' or '  '

	BLACKLISTED_REGEXES: List[str] = [\
		PHONE_NUMBER_REGEX,  # matches phone numbers
		URL_REGEX,  # matches URLs
		EMAIL_REGEX,  # matches email addresses
		NUMERIC_REGEX,  # matches numeric values
		GEO_REGEX,  # matches geographic coordinates
		LONG_REGEX,  # matches "long values, such as verbose descriptions"
		TIME_REGEX,
		DATE_REGEX,
		EMPTY_REGEX
	]  # => cf. Quercini & Reynaud "Entity Discovery and Annotation in Tables"

	BLACKLISTED_COMPILED_REGEXES =\
		[re.compile(regex_string) for regex_string in BLACKLISTED_REGEXES]


	@classmethod
	def column_is_blacklisted(cls, column: List[str],\
		min_black_list_matches: int = 1) -> bool:
		"""
		A helper function for get_identifying_column().
		Read documentation there for a detailed explanaition of the
		`min_black_list_matches` parameter.
		"""

		if min_black_list_matches > 0:
			return len([cell for cell in column\
				if any(regex.fullmatch(cell)\
				for regex in Table.BLACKLISTED_COMPILED_REGEXES)])\
				>= min_black_list_matches
		else:  # min_black_list_matches <= 0:
			return len([cell for cell in column\
				if any(regex.fullmatch(cell)\
				for regex in Table.BLACKLISTED_COMPILED_REGEXES)])\
				>= len(column) + min_black_list_matches


	@classmethod
	def column_uniqueness(cls, column: List[str]) -> float:
		"""
		A helper function for get_identifying_column().
		Returns a score between 0.0 and 1.0 signifying how "unique" a column
		is (the score is 1.0 iff all cells in the column are unique;
		for constant columns the score tends towards 0.0).
		"""
		return len(set(column)) / len(column)


	@classmethod
	def column_is_unique(cls, column: List[str],\
		min_uniqueness: float = 0.5) -> bool:
		"""
		A helper function for get_identifying_column().
		Returns whether a column consists of at least
		min_uniqueness*100% unique values
		(i.e. only for min_uniqueness=1.0, the column is actually required
		 to consist of all truly unique values).
		For min_uniqueness=0.0, this function always returns True.
		"""
		return len(set(column)) >= min_uniqueness * len(column)


	def get_identifying_column(self, min_uniqueness: float = 0.5,\
		consider_non_textual_columns_as_last_resort: bool = False,\
		min_black_list_matches: int = 1) -> List[str]:
		"""
		Get (the extension of) the identifying column of this table.

		Ritze et al. call this identifying column the "entity label attribute"
		in their "Matching HTML Tables to DBpedia" paper (T2K Match).
		They use the following heuristic to identify it:
		(a) take the string attribute with the highest number of unique values
		(b) use the left-most attribute in case of a tie

		Quercini & Reynaud's paper
		"Entity Discovery and Annotation in Tables" however provides us with
		yet another valuable idea:
		ignoring cells whose content matches the regular expression for URLs,
		email addresses, phone numbers, geographic coordinates, etc.

		We combine these two ideas like this:
		(1) Is there exactly one unique column, ignoring other unique columns
		    that match one of the "blacklisted" regexes?
		    => if so, return that one.
		    (the fact that it matches none of the regexes implies
		     that it is a textual column)
		(2) Are there multiple unique columns, ignoring other unique columns
		    that match one of the "blacklisted" regexes?
		    => if so, return the left-most of these.
		(3) Return the column with the highest number of unique values,
		    ignoring other columns that match one of the "blacklisted" regexes,
		    but only if at least 50% of the values are unique.
		    In case of a tie, use the left-most attribute.
		(4) Only if consider_non_textual_columns_as_last_resort=True
		    (by default, it's set to False):
		    If no column reaches the 50% uniqueness value, also consider
		    columns that match one of the "blacklisted" regexes, but keeping
		    the requirement of 50% uniqueness.
		(5) No candidate for an identifying column was found.
		    Return the empty list [].

		Note that the 50% threshhold can be altered using the `min_uniqueness`
		parameter (float value between 0.0 and 1.0, default value = 0.5).
		With a `min_uniqueness` of 0.0 and
		`consider_non_textual_columns_as_last_resort` set to True,
		there will always be an identifying column found in step (3),
		or (4) when all columns match a "blacklisted" regex.

		The `min_black_list_matches` parameter can be used to adjust when a
		column is considers to "match" a blacklisted regex:
		* min_black_list_matches=1 (default):
		  A column is considered a match, even if just one cell in it
		  matches a blacklisted regex.
		* min_black_list_matches=n:
		  A column is considered a match, if at least n cells in it
		  match a blacklisted regex.
		* min_black_list_matches=0:
		  A column is considered a match, only if **all** cells in it
		  match a blacklisted regex.
		* min_black_list_matches=-1:
		  A column is considered a match, only if all cells, except one, in it
		  match a blacklisted regex.
		* min_black_list_matches=-n:
		  A column is considered a match, only if all cells, except n,
		  i.e. at least len(column)-n, in it match a blacklisted regex.
		"""

		blacklisted_columns: List[List[str]] = []
		non_blacklisted_columns: List[List[str]] = []
		for col in self.columns:
			if Table.column_is_blacklisted(col, min_black_list_matches):
				blacklisted_columns.append(col)
			else:
				non_blacklisted_columns.append(col)

		# (1) Is there exactly one unique column (ignoring blacklisted)?
		#     => if so, return that one.
		#
		# (2) Are there multiple unique columns (ignoring blacklisted)?
		#     => if so, return the left-most of these.
		unique_non_blacklisted_columns =\
			[col for col in non_blacklisted_columns\
			if Table.column_is_unique(col, min_uniqueness=1.0)]
		if len(unique_non_blacklisted_columns) > 0:
			return unique_non_blacklisted_columns[0]  # (return left-most col)

		# (3) Return the column with the highest number of unique values
		#     (ignoring blacklisted and at least min_uniqueness*100% unique,
		#      using the left-most in case of a tie):
		most_unique_columns: List[Tuple[List[str], float]] =\
			sorted([(col, Table.column_uniqueness(col))\
			for col in non_blacklisted_columns],\
			key=lambda tuple: tuple[1], reverse=True)
		if len(most_unique_columns) > 0\
			and most_unique_columns[0][1] >= min_uniqueness:
			return most_unique_columns[0][0]

		# (4) Only if consider_non_textual_columns_as_last_resort=True:
		#     Like step (3) but now (including) blacklisted columns:
		if consider_non_textual_columns_as_last_resort:
			most_unique_columns: List[Tuple[List[str], float]] =\
			sorted([(col, Table.column_uniqueness(col))\
			for col in blacklisted_columns],\
			key=lambda tuple: tuple[1], reverse=True)
			if len(most_unique_columns) > 0\
				and most_unique_columns[0][1] >= min_uniqueness:
				return most_unique_columns[0][0]

		# (5) No candidate for an identifying column was found.
		#     => Return the empty list [].
		return []


	def classify(self,\
				 useSBERT:bool, useBing=False, useWebIsAdb=False,
				 useTextualSurroundings=True, textualSurroundingsWeighting=1.0,\
				 useAttrNames=True, attrNamesWeighting=1.0,\
				 useAttrExtensions=True, attrExtensionsWeighting=1.0,\
				 normalizeApproaches=False,
				 printProgressTo=sys.stdout)\
				-> List[Tuple[float, WikidataItem]]:
		"""
		Classify this table semantically.
		Returns an ordered list of WikidataItems that might represent the entity
		type of the tuples of this table, each with a floating-point score
		indicating how well it fits the table.
		
    	Keyword arguments:
    	useTextualSurroundings -- whether to use approach #1 if possible
    	                          (default True)
    	textualSurroundingsWeighting -- how to weight approach #1
    	                                (default 0.0)
    	useAttrNames -- whether to use approach #2 if possible
    	                (default True)
    	attrNamesWeighting -- how to weight approach #1
                              (default 0.0)
    	useAttrExtensions -- whether to use approach #3 if possible
    	                     (default True)
    	attrExtensionsWeighting -- how to weight approach #1
    	                           (default 0.0)
    	normalizeApproaches -- whether to normalize the result of each of the 3
    	                       approaches into the [0,1] range
    	                       (default False)

		More about `normalizeApproaches`:

    	Whether to normalize the rankings of classifyUsingTextualSurroundings(),
	    classifyUsingAttrNames(), classifyUsingAttrExtensions() into the same
	    range, i.e. the interval [0,1].
	 
	    When normalizeApproaches == False, the final ranking of classify()
	    is computed as:
	    classify() =
	        TEXTUAL_SURROUNDINGS_WEIGHTING * classifyUsingTextualSurroundings()
	      + ATTR_NAMES_WEIGHTING * classifyUsingAttrNames()
	      + ATTR_EXTENSIONS_WEIGHTING * classifyUsingAttrExtensions()
	 
	    When normalizeApproaches == True, the final ranking of classify()
	    is computed as:
	    classify() =
	        TXT_SURR_WGHT * normalize(classifyUsingTextualSurroundings())
	      + ATTR_NAMES_WEIGHTING * normalize(classifyUsingAttrNames())
	      + ATTR_EXTENSIONS_WEIGHTING * normalize(classifyUsingAttrExtensions())
	    => Advantage:
	       The weightings truly reflect the contribution of each approach.
	    => Disadvantage:
	       The information that one approach might be more confident than
	       another approach is lost.

	    `printProgressTo` may be set to `sys.stdout`, `sys.stderr` or
	    `open(os.devnull,"w")` for example.
    	"""

		resultUsingTextualSurroundings: Dict[WikidataItem, float] = {}
		if useTextualSurroundings and self.surroundingText != "":
			print("[PROGRESS] Classifying using textual surroundings...",\
				file=printProgressTo)
			resultUsingTextualSurroundings =\
				self.classifyUsingTextualSurroundings()
			if DEBUG:
				print("[DEBUG] Result using textual surroundings: " +\
					f"{debug_dict_sorted(resultUsingTextualSurroundings)}")

		resultUsingAttrNames: Dict[WikidataItem, float] = {}
		if useAttrNames and self.headerRow != []:
			print("[PROGRESS] Classifying using attribute names...",\
				file=printProgressTo)
			resultUsingAttrNames = self.classifyUsingAttrNames(\
				useSBERT=useSBERT)
			if DEBUG:
				print("[DEBUG] Result using attribute names: " +\
					f"{debug_dict_sorted(resultUsingAttrNames)}")


		resultUsingAttrExtensions: Dict[WikidataItem, float] = {}
		if useAttrExtensions and self.columns != []:
			print("[PROGRESS] Classifying using attribute extensions...",\
				file=printProgressTo)
			resultUsingAttrExtensions = self.classifyUsingAttrExtensions(\
				useBing=useBing, useWebIsAdb=useWebIsAdb)
			if DEBUG:
				print("[DEBUG] Result using attribute extensions: " +\
					f"{debug_dict_sorted(resultUsingAttrExtensions)}")

		if normalizeApproaches:
			resultUsingTextualSurroundings =\
				normalize(resultUsingTextualSurroundings)
			resultUsingAttrNames = normalize(resultUsingAttrNames)
			resultUsingAttrExtensions = normalize(resultUsingAttrExtensions)

		combinedResult: List[Tuple[float, WikidataItem]] = []

		combinedResult = combine3(\
			dct1=resultUsingTextualSurroundings,\
			weight1=textualSurroundingsWeighting,\
			dct2=resultUsingAttrNames,\
			weight2=attrNamesWeighting,\
			dct3=resultUsingAttrExtensions,\
			weight3=attrExtensionsWeighting,\
		)

		return combinedResult

	def classifyUsingTextualSurroundings(self) -> Dict[WikidataItem, float]:
		"""
		Classify this table using the "Using Textual Surroundings" approach,
		implemented in filter_nouns_with_heuristics.py.
		Higher scores mean better matches.
		"""

		return filter_nouns_with_heuristics_as_dict(\
			input_text=self.surroundingText,\
			VERBOSE=False,\
			ALLOW_EQUAL_SCORES=True)
		# (already returns a Dict[WikidataItem, float])

	def classifyUsingAttrNames(self, useSBERT: bool)\
		-> Dict[WikidataItem, float]:
		"""
		Classify this table using the "Using Attribute Names" approach,
		implemented in attr_names_to_ontology_class.py.
		Higher scores mean better matches.

		When useSBERT==False, the Jaccard index is used for computing
		word similarity instead.
		"""

		# First, clean the header row / attribute names:
		# (a) remove empty column captions ("")
		# (b) remove column captions shorter than 3 characters when using
		#     Jaccard-trigrams instead of SBERT
		cleaned_headerRow: List[str] =\
			[header for header in self.headerRow\
			 if header.strip() != "" and (len(header) >= 3 or useSBERT)]

		# attr_names_to_ontology_class() returns a Dict[str, float]
		# (with the strings being names of DBpedia classes)
		# but we need a Dict[WikidataItem, float]:
		return { WikidataItem(\
			get_dbpedia_classes_mapped_to_wikidata()[dbpediaClass]) : score\
			for dbpediaClass, score\
			in attr_names_to_ontology_class(\
				inputAttrNames=cleaned_headerRow,
				USE_BETTER_SUM_FORMULA=True,
				USE_SBERT_INSTEAD_OF_JACCARD=useSBERT,
				VERBOSE=False
			).items()}

	def classifyUsingAttrExtensions(self, useBing=False, useWebIsAdb=False)\
		 -> Dict[WikidataItem, float]:
		"""
		Classify this table using the "Using Attribute Extensions" approach,
		implemented in attr_extension_to_ontology_class.py.
		Higher scores mean better matches.

		The `useBing` and `useWebIsAdb` parameters
		shall not both be set to True!
		"""

		if useBing:
			return\
				attr_extension_to_ontology_class_web_search_list_onto_as_dict(\
				cell_labels=self.get_identifying_column())
			# (already returns a Dict[WikidataItem, float])
		elif useWebIsAdb:
			exit("[ERROR] Using the WebIsA DB is not yet implemented!")
			return None  # (possible ToDo: not even coded yet...)
		else:
			return\
				attr_extension_to_ontology_class(\
				cell_labels=self.get_identifying_column())
			# (returns a Dict[WikidataItem, int])


	@classmethod
	def identifyCSVdialect(cls, csv_str: str) -> Dialect:
		return csv.Sniffer().sniff(csv_str)

	@classmethod
	def parseCSV(cls, csv_str: str, dialect: Dialect = None) -> Table:
		columns: List[List[str]] = []

		# Documentation of
		#   csv.reader(csvfile, dialect='excel', **fmtparams):
		# "csvfile can be any object which supports the iterator protocol
		#  and returns a string each time its __next__() method is called —
		#  file objects and list objects are both suitable. If csvfile is a
		#  file object, it should be opened with newline=''."

		csv_reader = csv.reader(csv_str.splitlines(keepends=True), dialect\
			if dialect is not None else Table.identifyCSVdialect(csv_str))
		for row in csv_reader:
			no_of_cols: int = len(row)  # (number of columns)
			if columns == []:
				columns = [[] for i in range(0, no_of_cols)]  # (initialize)
			for col_index in range(0, no_of_cols):
				columns[col_index].append(row[col_index])

		result = Table(surroundingText="", headerRow=[],\
			columns=columns)  # surroundingText is not applicable to CSV files
		result.parse_header_row()  # set headerRow!=[] if possible
		return result

	@classmethod
	def parseXLSX(cls, xlsxPath: str) -> Table:
		# Use openpyxl to parse Microsoft Excel files,
		# cf. https://www.geeksforgeeks.org/
		#   working-with-excel-spreadsheets-in-python/

		import openpyxl
		wb_obj = openpyxl.load_workbook(xlsxPath)
		sheet_obj = wb_obj.active

		# # Example from www.geeksforgeeks.org :
		# cell_obj = sheet_obj.cell(row = 1, column = 1)
		# print(cell_obj.value)

		# # Works but does not consider a possible header row:
		# columns = [list(map(lambda cell_obj: str(cell_obj.value), column))\
		# 	for column in sheet_obj.iter_cols()]

		rows = [list(map(lambda cell_obj: str(cell_obj.value), row))\
			for row in sheet_obj.iter_rows()]

		#If Table.parseCSV() took an iterator instead, it would look like this:
		#return Table.parseCSV(csv_lines=(",".join(row) + "\n" for row in rows))
		csv_str = "\n".join([",".join(row) for row in rows])
		return Table.parseCSV(csv_str)

	@classmethod
	def parseJSON(cls, json_str: str, onlyRelational=False,\
		onlyWithHeader=True, useAdditionalHeuristics=False) -> Optional[Table]: # ToDo: change to onlyRelational=True after testing!
		"""
		Parses a .JSON file of the format that's used in the
		WDC Web Table Corpus (http://webdatacommons.org/webtables/2015/
		downloadInstructions.html):

		{"relation":[[..., ...],[..., ...]],
		 "pageTitle":"...",
		 "title":"...",
		 "url":"https://...",
		 "hasHeader":true,
		 "headerPosition":"MIXED",
		 "tableType":"RELATION",
		 "tableNum":6,
		 "s3Link":"common-crawl/crawl-data/...",
		 "recordEndOffset":867273895,
		 "recordOffset":867263717,
		 "tableOrientation":"HORIZONTAL",
		 "textBeforeTable":"...",
		 "textAfterTable":"...",
		 "hasKeyColumn":true,
		 "keyColumnIndex":0,
		 "headerRowIndex":0}

		This method returns None for invalid JSON inputs
		(i.e. not of the above format), for non-relational tables
		(if onlyRelational=True) and for tables without a header
		(if onlyWithHeader=True)!
		"""
		try:
			j = json.loads(json_str)
			
			if onlyRelational and j["tableType"] != "RELATION":
				# e.g. "tableType":"ENTITY"
				return None  # skip non-relational tables

			surroundingText: str =\
				j["textBeforeTable"] + " " + j["textAfterTable"]
			headerRow: List[str] = None
			columns: List[List[str]] = None

			# For the header information, rely on the information given
			#   (i.e. do no own heuristics):
			if j["hasHeader"]:
				if j["headerPosition"] == "FIRST_ROW":  # (most common case)
					headerRow = [col[0] for col in j["relation"]]
					columns = [col[1:] for col in j["relation"]] # strip header
				elif j["headerPosition"] == "FIRST_COLUMN":
					# Tranpose the table because for us, the header is always
					#   in the first row:
					headerRow = j["relation"][0]  # 1st column = header
					columns = transpose(j["relation"][1:]).tolist()
				elif j["headerPosition"] == "MIXED":
					return None  # skipping such unsure/non-relational? tables
				else:
					print("[PARSE ERROR] Unknown value for 'headerPosition'" +\
						f""" in JSON: '{j["headerPosition"]}'""",\
						file=sys.stderr)
					return None
			else:  # "hasHeader":false => "headerPosition":"NONE"
				if onlyWithHeader and not useAdditionalHeuristics:
					# Corpus says that this table has no header, we shall
					#   not use additional heuristcs and we shall only
					#   return tables with header => return None:
					return None
				elif not onlyWithHeader and not useAdditionalHeuristics:
					# Corpus says that this table has no header, we **are**
					#   allowed to return tables without a header but we are
					#   **not** allowed to use additional heuristics
					#   => return the table without header
					#      as it is in the corpus:
					headerRow = []
					# We still have to find out which way to orient the columns
					#   though:
					columns = None  # ToDo(!)
				elif useAdditionalHeuristics:
					# We **are** allowed to use additional heuristics:
					# Looking at some examples shows that when
					#   "hasHeader":false, there **is** usually a header
					#   in the first column...
					return None  # ToDo(!)

			return Table(surroundingText=surroundingText, headerRow=headerRow,\
				columns=columns)
		except BaseException as e:
			# (most likely a KeyError when JSON is in invalid format)
			if DEBUG:
				print(f"[DEBUG] Error parsing JSON: {e}", file=sys.stderr)
			return None

	@classmethod
	def parseTAR(cls, tarPath: str, csv_dialect: Dialect = None)\
		-> Iterator[Table]:
		# cf. https://docs.python.org/3/library/tarfile.html
		tar = tarfile.open(tarPath)
		for tarinfo in tar:
			# Skip directories and hidden files in the TAR archive:
			if tarinfo.isfile() and tarinfo.name[:1] != ".":
				if os.path.splitext(tarinfo.name)[1] == ".json":
					yield Table.parseJSON(\
						json_str=tar.extractfile(tarinfo).read())
				elif os.path.splitext(tarinfo.name)[1] == ".csv":
					yield Table.parseCSV(\
						csv_str=tar.extractfile(tarinfo).read(),\
						dialect=csv_dialect)
				elif os.path.splitext(tarinfo.name)[1] in [".xlsx", ".xls"]:
					# Because parseXLSX() expects a path, we have to
					#   temporarily extract the .xlsx file to the file system:
					table: Table = None
					with tempfile.NamedTemporaryFile() as temp_path:
						# Extract the .xlsx file to a temporary location:
						tar.extract(tarinfo, path=temp_path)
						# Parse that temporary .xlsx file:
						table = Table.parseXLSX(xlsxPath=temp_path)
					yield table
				# All other file types in the TAR archive are ignored.


	@classmethod
	def parseFolder(cls, folderPath: str, recursive=True,\
		csv_dialect: Dialect = None) -> Iterator[Table]:
		# Sorting is important such that the order is predictable to the user:
		for folder_item in sorted(os.listdir(folderPath)):
			folder_item = os.path.join(folderPath, folder_item)
			if os.path.isfile(folder_item):
				file_name, file_extension = os.path.splitext(folder_item)
				if file_extension == ".csv":
					with open(folder_item, 'r', newline='') as csv_file:
						# (newline='' is required by the CSV library)
						yield Table.parseCSV(csv_str=csv_file.read(),\
							dialect=csv_dialect)
				elif file_extension in [".xlsx", ".xls"]:
					yield Table.parseXLSX(xlsxPath=folder_item)
				elif file_extension == ".json":
					with open(folder_item, 'r') as json_file: 
						yield Table.parseJSON(json_str=json_file.read())
				elif file_extension == ".tar":
					for table in Table.parseTAR(tarPath=folder_item):
						yield table
				# Skip files with any other filetype/extension.
			elif recursive:
				for table in Table.parseFolder(folderPath=folder_item,\
					recursive=True):
					yield table
			# When the item is a directory and recursive=False, skip it.

	@classmethod
	def parseCorpus(cls, corpusPath: str, recursive=True,\
		yieldNones=False, csv_dialect: Dialect = None) -> Iterator[Table]:
		if os.path.isfile(corpusPath):
			file_name, file_extension = os.path.splitext(corpusPath)
			if file_extension == ".tar":
				return filter(lambda table: yieldNones or table is not None,\
					Table.parseTAR(tarPath=corpusPath, csv_dialect=csv_dialect))
			else:
				sys.exit("Error: --corpus supplied either has to be a "+\
					" directory or a .TAR file.")
		else:
			return filter(lambda table: yieldNones or table is not None,\
				Table.parseFolder(folderPath=corpusPath, recursive=recursive,\
					csv_dialect=csv_dialect))


def main():
	parser = argparse.ArgumentParser(
		description="""NETT - Narrative Entity Type(s) to Tables.
		NETT can be run in 4 modes:
		(1) --stats, no ENTITY_TYPE(s) supplied => statistics for given corpus
		(2) --stats, ENTITY_TYPE(s) supplied => statistics for given corpus,
		with focus on the supplied entity types (specifying additional
		narrative knowledge is possible)
		(3) no ENTITY_TYPE(s) supplied => classify all tables in given corpus
		(but max. 10,000 tables)
		(4) ENTITY_TYPE(s) supplied => main productive mode: search corpus
        for tables based on given entity/-ies
        (and possibly narrative knowledge)!!
		""")
	# ToDo: list a few example calls in this description!!

	parser.add_argument(
    	'entityTypes', # Narratives = "Knowing what to look for"
    	type=str,
    	help="""A list of entity types (the entity types to look for).
    	Ideally they're Wikidata ID's of the form 'Q000000'.
    	Alternatively they're names of DBpedia classes which are then
    	mapped to Wikidata ID's using a pre-programmed mapping table.
    	When they are neither of the form 'Q000000' nor the name of a DBpedia
    	class, Wikidata is searched with them as a search string and the user
    	asked which of the search results is the one they meant.
    	When no entity types are supplied at all, a ranked list of entity types
    	is returned for every table in the corpus instead (mode (3))
    	(assuming that the --stats flag isn't supplied).""",
    	nargs='*',
    	metavar='ENTITY_TYPE')

	parser.add_argument('--corpus',
    	type=str,
    	default='',
    	help="""Path to a folder containing tables as CSV/JSON/TAR files.
    	Or path to a single TAR file containing tables as CSV/JSON files.
    	You may use 'default_corpus' for testing
    	(only contains a handful of tables!).
    	Note that all tables smaller than 3x3 are rigorously filtered out.
    	Folders are parsed in alphabetical order.
    	Excel files instead of CSV files are supported too when the
    	`openpyxl` Python module is installed:
    	`pip install openpyxl` or
    	`python3 -m pip install openpyxl`""",
    	metavar='CORPUS_PATH',
    	required=True)

	parser.add_argument('--corpus-non-recursive',
		action='store_true',
		help="""When a folder is specified for the --corpus parameter, do NOT
		look into its subfolders recursively.""")

	parser.add_argument('--stats',
		action='store_true',
		help="""Ask the user to classify each table in the corpus and return
		statistics in the end. When entity types are supplied as well,
		only tables are shown to the user where NETT believes they *might*
		be of one of these entity types. The statictics in the end are then
		specific to these entity types.""")

	parser.add_argument('-k',
    	type=int,
    	default=1,
    	help="""Only return tables for which the entity type searched for was
    	in the top-k results. By default, k=1 which means that only tables are
    	returned for which the entity type searched for was the best guess.""",
    	metavar='K')

	parser.add_argument('--threshold',
    	type=float,
    	default=0.0,
    	help="""Only return tables for which the entity type searched for got
    	a score of at least this threshold (floating-point number).
    	This makes sense particularly when used together with the --normalize
    	flag. When --normalize is set and all weights are kept at their default
    	value of 1.0, then this value shall be somewhere between 0.0 and 3.0.
    	This is a filter that is set in addition to the -k filter!
    	By default, this threshold is set to 0.0, i.e. it has no effect.""",
    	metavar='THRESHOLD')

	parser.add_argument('--dont-use-textual-surroundings',
		action='store_true',
		help="""Do not take textual surroundings of tables into account.
		This only has an effect for JSON input.""")

	parser.add_argument('--dont-use-attr-names',
		action='store_true',
		help='Do not take the attribute/column names of tables into account.')

	parser.add_argument('--dont-use-attr-extensions',
		action='store_true',
		help='Do not take attribute extensions of tables into account.')

	parser.add_argument('--prefer-textual-surroundings',
		action='store_true',
		help="""Use textual surroundings only, unless this approach yields no
		results whatsoever, then use the other ones.""")

	parser.add_argument('--prefer-attr-names',
		action='store_true',
		help="""Use attribute/column names only, unless this approach yields no
		results whatsoever, then use the other ones.""")

	parser.add_argument('--prefer-attr-extensions',
		action='store_true',
		help="""Use attribute extensions only, unless this approach yields no
		results whatsoever, then use the other ones.""")

	parser.add_argument('--require-textual-surroundings',
		action='store_true',
		help="""Skip tables without textual surroundings entirely.
		This has the effect of only considering JSON files in the corpus.""")

	parser.add_argument('--require-attr-names',
		action='store_true',
		help='Skip tables without attribute names entirely.')

	parser.add_argument('--require-attr-extensions',
		action='store_true',
		help='Skip tables without an identifying/unique column entirely.')

	parser.add_argument('--normalize',
		action='store_true',
		help="""Whether to normalize the result of each of the 3
    	approaches into the [0,1] range""")

	parser.add_argument('--textual-surroundings-weight',
    	type=float,
    	default=1.0,
    	help='How to weight the textual surroundings approach.',
    	metavar='WEIGHTING')

	parser.add_argument('--attr-names-weight',
    	type=float,
    	default=1.0,
    	help='How to weight the attribute names approach.',
    	metavar='WEIGHTING')

	parser.add_argument('--attr-extensions-weight',
    	type=float,
    	default=1.0,
    	help='How to weight the attribute extensions approach.',
    	metavar='WEIGHTING')

	# (Advanced) ToDo: if all weights are set to 0.0 try out ("learn") which
	#   weights lead to the best results!

	parser.add_argument('--csv-delimiter',
    	type=str,
    	default='',
    	help="""
    	Specify the delimiter character (e.g. ',' or ';') that's used by the
    	CSV files in the corpus.
    	By default, it is automatically recognized for each CSV file
    	individually. So specify this only when all CSV files in the corpus
    	have a common format!
    	""",
    	metavar='CSV_DELIMITER')

	parser.add_argument('--csv-quotechar',
    	type=str,
    	default='',
    	help="""
    	Specify the quotechar character (e.g. '"') that's used by the
    	CSV files in the corpus.
    	By default, it is automatically recognized for each CSV file
    	individually. So specify this only when all CSV files in the corpus
    	have a common format! 
    	""",
    	metavar='CSV_QUOTECHAR')

	parser.add_argument('--ordered',
		action='store_true',
		help="""When this flag is set, the tables in the output are sorted by
		how well they score. Beware that lazy output is lost which means that
		this flag should not be used together with a very big corpus!
		This flag only has an effect when in "productive mode" (4), i.e. when
		entity types are supplied and --stats is *not* set.""")

	parser.add_argument('--bing',
		action='store_true',
		help="""
		Use the Bing web search and Wikidata for inferring the semantics of
		attribute extensions instead of just Wikidata.
		For this to work, the BING_SEARCH_V7_SUBSCRIPTION_KEY and
		BING_SEARCH_V7_ENDPOINT variables need to be set. You can find
		their values on portal.azure.com
		""")

	parser.add_argument('--webisadb',
		action='store_true',
		help="""
		Use the WebIsA Database from webdatacommons.org/isadb and Wikidata for
		inferring the semantics of attribute extensions instead of just
		Wikidata (cf. the --bing flag which contradicts this flag).
		For this to work, the corresponding MongoDB has to be running
		on localhost.
		THIS FEATURE IS NOT IMPLEMENTED!
		""")

	parser.add_argument('--verbose', '-v',
		action='store_true',
		help='Print verbose info prints.')

	parser.add_argument('--debug',
		action='store_true',
		help='Print debug info prints.')

	# cf. https://stackoverflow.com/questions/7869345/
	#     how-to-make-python-argparse-mutually-exclusive-
	#     group-arguments-without-prefix:
	group = parser.add_mutually_exclusive_group()
	group.add_argument('--jaccard', action='store_true',
		help="Use Jaccard index for word similarity of column/attribute names.")
	group.add_argument('--sbert', action='store_true',
		help="""
		Use SBERT for word similarity of column/attribute names.
		This is currently slower than using the Jaccard index.
		For this to work you need to run `pip install -U sentence-transformers`
		or `python3 -m pip install -U sentence-transformers` first!
		""")

	# Narratives = "Taking Advantage of the Knowledge in the Narrative":

	parser.add_argument('--co-occurring-keywords',
    	type=str,
    	default='',
    	help="""
    	List other words occurring in the narrative. The results are then
    	limited to those tables where at least one of these words occurs in
    	either the surrounding text of the table or inside the table.
    	""",
    	nargs='*',
    	metavar='KEYWORD')

	parser.add_argument('--co-occurring-keywords-all',
		action='store_true',
		help="""
		Set this flag to require all(!) keywords specified with the 
		--co-occurring-keywords parameter to occur in the surrounding text
		of or inside each table.
		(By default, one keyword occuring is regarded as being sufficient.)
		""")

	parser.add_argument('--attribute-cond',
    	type=str,
    	default='',
    	help="""
    	List one or multiple attribute conditions of the form
    	"[attribute name] [<=,>=,<,>,==,!=,in,not in]
    	 [value, value range or value list]", e.g.
        "horsepower >= 500" or
        "year in range(1980,2000)" or
        "firstName not in ['Alex', 'Alexander']"
    	When supplied, the results are limited to those tables where
    	(a) a column name could be matched to each specified attribute name
    	   (--jaccard or --sbert is used, depending on parameter)
    	and
    	(b) all(!) values in the extension of that column fulfill
    	    the specified condition.
    	Note that (a) and (b) are combined, i.e. columns where all values
    	fulfill the condition in (b) are considered to be more likely to match
    	the attribute in (a).
    	Use the --attribute-cond-strictness parameter
    	to specify the strictness of
    	this condition check as a value between 0.0 (no check at all)
    	and 1.0 (100-percent strict).
    	""",
    	nargs='*',
    	metavar='CONDITION')

	parser.add_argument('--attribute-cond-strictness',
    	type=float,
    	default=1.0,
    	help="""
    	Only meaningful in combination with the --attribute-cond parameter.
    	A value between 0.0 (least strict) and 1.0 (most strict, default).
    	""",
    	metavar='STRICTNESS')

	args = parser.parse_args()

	corpus: str = args.corpus
	stats: bool = args.stats
	entityTypes: List[str] = args.entityTypes

	DEBUG = args.debug

	# <preparation>
	print("[PREPARING] Mapping DBpedia properties to SBERT vectors...")
	initalize_dbpedia_properties_mapped_to_SBERT_vector()
	print("[PREPARING] Done.")
	# </preparation>

	if stats and entityTypes == []:
		# (1) Corpus supplied, statistics requested (evaluation feature):
		#   * Program has to ask user (with the help of pretty_print())
		#     for every table which mapping is correct.
		#   * The user enters enters "1", "2", "3", ..., "X", "N/A" for each
		#     table presented, until the whole corpus has been annotated or
		#     until the user stops by entering "finish".
		#   * Then, the statistics (MRR, Top-k coverage; for using
		#     1, 2 or 3 of the three approaches and for various
		#     weightings and preferences of these approaches; also specifically
		#     the "k-recall" for all(!) entity types annotated) are printed
		#     and the user is asked whether they want to export
		#     their annotations (Y/n).
		#   * The --co-occurring-keywords and --attribute-cond parameters
		#     (the "narrative parameters") are ignored in this case as they
		#     make no sense when no entity types are specified.
		print("This combination of parameters is not yet implemented.")  # ToDo!
	elif stats and entityTypes != []:
		# (2) Corpus and entity types supplied, statistics requested
		#     (evaluation feature):
		#   * Like the case (1) but this time, looking for one (or multiple)
		#     specific entity types.
		#   * Tables for which NETT believes their tuples definitely do not
		#     belong to any of these entity types, are skipped and not
		#     presented to the user at all.
		#   * Tables for which NETT believes their tuples might belong to any
		#     of these entity types, are presented to the user in the same way
		#     as in case (1).
		#   * In the end, statistics are printed (just as in case (1));
		#     beware however that the statistics may not be entirely accurate
		#     when NETT wrongly skipped tables it thought definitely do not
		#     represent any of the entity types searched for.
		#   * Note that it also makes sense to use the --co-occurring-keywords
		#     and --attribute-cond parameters here
		#     (the "narrative parameters").
		#   * The user is shown both tables that match this narrative knowledge
		#     and tables that don't (for a better evaluation afterwards).
		print("This combination of parameters is not yet implemented.")  # ToDo
	elif not stats and entityTypes == []:
		# (3) Corpus supplied, entity-type-mappings requested
		#     (evaluation feature):
		#   * Map all tables of the given corpus to the top-k entities.
		#   * It might be sensible to change k to a bigger value
		#     than 1 (default).
		#   * For very big corpora, only the first 10,000 annotatable
		#     relational tables are considered. Then, the program terminates.
		#   * The --co-occurring-keywords and --attribute-cond parameters
		#     (the "narrative parameters") are ignored in this case as they
		#     make no sense when no entity types are specified.
		print("This combination of parameters is not yet implemented.")  # ToDo
	elif not stats and entityTypes != []:
		# (4) Corpus and entity types supplied, tables requested
		#     (the main productive feature!!!):
		#   * Search the corpus for tables whose tuples represent one
		#     of the given entity types, possibly making use of the narrative
		#     knowledge provided with the --co-occurring-keywords and
		#     --attribute-cond parameters (the "narrative parameters").
		#   * Depending on your requirements for precision and recall, you may
		#     want to change the -k parameter to a value other than 1.
		#     With k=1, only tables are retured for which the entity type
		#     searched for was the best match
		#     (i.e. highest possible precision, lowest possible recall).
		#     For bigger k, the recall is higher but the precision lower.
		print("This combination of parameters is not yet implemented.")  # ToDo!


if __name__ == "__main__":
	main()
