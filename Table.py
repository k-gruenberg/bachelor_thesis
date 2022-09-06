from __future__ import annotations
from typing import List, Dict, Any, Iterator, Set
import json  # https://docs.python.org/3/library/json.html
import re
import os
import sys
import csv  # https://docs.python.org/3/library/csv.html
import tarfile  # https://docs.python.org/3/library/tarfile.html
import tempfile  # https://docs.python.org/3/library/tempfile.html

from numpy import transpose

from WikidataItem import WikidataItem
from ClassificationResult import ClassificationResult

# Import the fundamental approaches
#     #1 "Using Textual Surroundings"
#     #2 "Using Attribute Names"
#     #3 "Using Attribute Extensions"
# from the respective Python files:
from filter_nouns_with_heuristics import filter_nouns_with_heuristics_as_dict
from attr_names_to_ontology_class import attr_names_to_ontology_class
from attr_extension_to_ontology_class import attr_extension_to_ontology_class
from attr_extension_to_ontology_class_web_search import\
	attr_extension_to_ontology_class_web_search_list_onto_as_dict

from nett_map_dbpedia_classes_to_wikidata import\
	get_dbpedia_classes_mapped_to_wikidata


def debug_dict_sorted(d: Dict[WikidataItem, float]) -> str:
	"""
	A debug-printable short string representation of a
	Dict[WikidataItem, float] dictionary, where the floats are to be
	interpreted as match scores (bigger score = better match).
	"""
	if len(d) <= 2:  # empy dictionary or dictionary with <= 2 key value pairs:
		return str(d)
	else:
		max_key: WikidataItem = None
		max_value: float = float('-inf')
		for key in d:
			value: float = d[key]
			if value > max_value:
				max_key = key
				max_value = value
		return "{" + str(max_key) + ": " + str(max_value) + ", ...}"


class Table:
	def __init__(self, surroundingText: str, headerRow: List[str],\
				 columns: List[List[str]]):
		"""
		Initalize a new Table that has already been parsed.

		If the `columns` haven't been parsed for a `headerRow` yet,
		set headerRow=[] and
		call the parse_header_row() method after construction!

		`headerRow` may be `[]` but never `None`!
		"""

		self.surroundingText = surroundingText  # (1) Using Textual Surroundings
		self.headerRow = headerRow  # (2) Using Attribute Names
		self.columns = columns  # (3) Using Attribute Extensions
		self.file_name = "???"

	def __eq__(self, other):
		if isinstance(other, Table):
			# Equality check should only matter when having parsed a file with
			#   existing annotations (--annotations-file parameter).
			# In that case, the tables should be exactly equal, so we can be
			#   this strict:
			return self.surroundingText == other.surroundingText\
				and self.headerRow == other.headerRow\
				and self.columns == other.columns\
				and self.file_name == other.file_name
		return False

	def width(self) -> int:
		"""
		Returns the width of this Table, i.e. the number of columns.
		"""
		return len(self.columns)  # width == # of columns

	def min_height(self, includingHeaderRow=True) -> int:
		"""
		Returns the height of this Table, i.e. the number of rows.
		Including the header row by default (if this Table has one, that is).
		Set includingHeaderRow=False to exclude a possible header row.

		The value returned is different to that of max_height() only if
		the columns don't all have the same height, which they should have.
		"""
		return min(len(col) for col in self.columns) +\
			(1 if includingHeaderRow and self.headerRow != [] else 0)

	def max_height(self, includingHeaderRow=True) -> int:
		"""
		Returns the height of this Table, i.e. the number of rows.
		Including the header row by default (if this Table has one, that is).
		Set includingHeaderRow=False to exclude a possible header row.

		The value returned is different to that of min_height() only if
		the columns don't all have the same height, which they should have.
		"""
		return max(len(col) for col in self.columns) +\
			(1 if includingHeaderRow and self.headerRow != [] else 0)

	def min_dimension(self, includingHeaderRow=True) -> int:  # ToDo: USE to filter out <3x3 tables!!
		"""
		Example: when this is a 4x10 table, returns 4.
		"""
		return min(self.width(),\
			self.min_height(includingHeaderRow=includingHeaderRow))

	def pretty_print(self, maxNumberOfTuples=6, maxColWidth=25,\
		maxTotalWidth=180) -> str:
		"""
		A pretty printable version of this Table, e.g.:

		Restaurant Name       | Rating | Price | Reviews
		------------------------------------------------
		Aquarius              |        | $$    | 0
		BIG & little's        |        | $     | 0
		Brown Bag Seafood Co. |        | $$    | 0
		...                   | ...    | ...   | ...

		(in this example, maxNumberOfTuples=3)
		"""

        # The width (number of characters) of each column, only regarding
        #   the first `maxNumberOfTuples` rows each:
		columnWidths: List[int] = list(map(\
			lambda column: max(map(\
				lambda cell: len(cell),\
				column[0:maxNumberOfTuples]\
			)),\
			self.columns\
			))
		# If a header name is longer than corresponding column width, it
		#   has to be increased further:
		if self.headerRow != []:
			columnWidths = list(\
				map(\
					lambda tuple: max(tuple[0], tuple[1]),\
					zip(columnWidths, map(lambda header: len(header),\
						self.headerRow))\
					)\
				)
		# Every column shall have a width of at least 3:
		columnWidths = [max(3, cw) for cw in columnWidths]
		# Every column shall have a width of at most `maxColWidth`:
		columnWidths = [min(maxColWidth, cw) for cw in columnWidths]

		def print_row(cells: List[str], widths: List[int]) -> str:
			res: str = ""
			for i in range(0, len(cells)):
				width = widths[i]
				res += cells[i][:width] + " " * (width-len(cells[i]))
				if i != len(cells)-1: res += " | "
			res += "\n"
			return res

		result: str = ""

		if self.headerRow != []:
			result += print_row(self.headerRow, columnWidths)
		else:  # no header row to print => print empty header row:
			result += print_row([""] * len(columnWidths), columnWidths)

		totalWidth = sum(columnWidths) + 3*(len(columnWidths)-1)
		result += "-" * totalWidth + "\n" # "-----------------------------"

		for y in range(0, min(maxNumberOfTuples, len(self.columns[0]))):
			result += print_row(\
				[self.columns[x][y] for x in range(0, len(self.columns))],\
				 columnWidths)

		if maxNumberOfTuples < len(self.columns[0]):  # print "... | ..." row:
			result += print_row(["..."] * len(columnWidths), columnWidths)

		if maxTotalWidth > 0:
			result = "\n".join(map(lambda line: line\
				if len(line) <= maxTotalWidth\
				else line[:maxTotalWidth-4] + " ...", result.splitlines()))

		return result


	# (possible ToDo: quoting?!)
	def as_csv(self, with_header=True, max_number_of_rows: int = -1) -> str:
		self_min_height: int = self.min_height(includingHeaderRow=False)

		if max_number_of_rows == -1:
			max_number_of_rows = self_min_height

		self_as_csv: str = ""

		if with_header and self.headerRow != []:
			self_as_csv += ",".join(\
				self.headerRow[col_index]\
				for col_index in range(0, len(self.headerRow))) + "\n"

		for row_index in range(0, min(max_number_of_rows, self_min_height)):
			self_as_csv += ",".join(\
				self.columns[col_index][row_index]\
				for col_index in range(0, self.width())) + "\n"

		return self_as_csv

	def parse_header_row(self) -> bool:
		"""
		When self.headerRow == []:
		    Tries to to identify a header row in self.columns using the
		    heuristic used by the csv.Sniffer().has_header() method
		    (cf. https://docs.python.org/3/library/csv.html).
		    If found, the header row is deleted from self.columns and
		    self.headerRow is set instead.
		    Returns True when a header row was found and False otherwise.
		When self.headerRow != []:
		    This method does nothing and returns True.

		Important note:
		When using one of the parseXXX() class methods,
		calling this function is NOT necessary!!
		"""
		if self.headerRow == []:
			self_as_csv_sample: str =\
				self.as_csv(with_header=False, max_number_of_rows=20)
			# From https://docs.python.org/3/library/csv.html,
			# regarding csv.Sniffer.has_header(sample):
			# "Twenty rows after the first row are sampled; if more than half
			#  of columns + rows meet the criteria, True is returned."
			if csv.Sniffer().has_header(self_as_csv_sample):
				# Set the header row to the first row:
				self.headerRow = [self.columns[col_index][0] for\
					col_index in range(0, len(self.columns))]
				# Delete the header row from self.columns where it
				#   does not belong:
				for col_index in range(0, len(self.columns)):
					del self.columns[col_index][0]
				return True
			else:
				return False  # csv.Sniffer().has_header() found no header.
		else:
			return True  # There already is a header row (parsed).


	# Correctly regex-matching URLs and email addresses is actually much
	# harder than one might think, read:
	#
	# Email addresses:
	# * https://stackoverflow.com/questions/9238640/
	#     how-long-can-a-tld-possibly-be
	# * https://stackoverflow.com/questions/201323/
	#     how-can-i-validate-an-email-address-using-a-regular-expression
	# * http://www.ex-parrot.com/~pdw/Mail-RFC822-Address.html
	# * https://en.wikipedia.org/wiki/Email_address#Local-part
	# * http://emailregex.com/
	#
	# URLs:
	# * https://stackoverflow.com/questions/3809401/
	#     what-is-a-good-regular-expression-to-match-a-url
	# * https://mathiasbynens.be/demo/url-regex:

	PHONE_NUMBER_REGEX = r"\+?\d[\d -]+\d"
	URL_REGEX = r"^(https:|http:|www\.)\S*"  # https://regex101.com/r/S2CbwM/1
	EMAIL_REGEX = r".+@.+"
	#EMAIL_REGEX = r".+@.+\..+"
	# -> problem: "jsmith@[IPv6:2001:db8::1]" is a valid email without any dots!
	#EMAIL_REGEX = r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)"
	# -> source: emailregex.com
	# -> problem: !#$%&'*+-/=?^_`{|}~ are allowed characters in emails too!
	NUMERIC_REGEX = r"[\+\-]?[\d., _]+"  # r"[\d.,]+"
	GEO_REGEX = r"""[\+\-\d.,;'" NESW]+"""
	LONG_REGEX = r".{50,}"
	TIME_REGEX = r"\d\d:\d\d(:\d\d)?"  # e.g. '00:15' or '00:15:00'
	DATE_REGEX = r"(\d\d? \w{3,} \d\d\d\d)|(\w{3,} \d\d?, \d\d\d\d)"
	# e.g. '25 Jan 1893' or 'August 31, 2006'
	EMPTY_REGEX = r"\s*"  # e.g. '' or ' ' or '  '

	BLACKLISTED_REGEXES: List[str] = [\
		PHONE_NUMBER_REGEX,  # matches phone numbers
		URL_REGEX,  # matches URLs
		EMAIL_REGEX,  # matches email addresses
		NUMERIC_REGEX,  # matches numeric values
		GEO_REGEX,  # matches geographic coordinates
		LONG_REGEX,  # matches "long values, such as verbose descriptions"
		TIME_REGEX,
		DATE_REGEX,
		EMPTY_REGEX
	]  # => cf. Quercini & Reynaud "Entity Discovery and Annotation in Tables"

	BLACKLISTED_COMPILED_REGEXES =\
		[re.compile(regex_string) for regex_string in BLACKLISTED_REGEXES]


	@classmethod
	def column_is_blacklisted(cls, column: List[str],\
		min_black_list_matches: int = 1) -> bool:
		"""
		A helper function for get_identifying_column().
		Read documentation there for a detailed explanaition of the
		`min_black_list_matches` parameter.
		"""

		if min_black_list_matches > 0:
			return len([cell for cell in column\
				if any(regex.fullmatch(cell)\
				for regex in Table.BLACKLISTED_COMPILED_REGEXES)])\
				>= min_black_list_matches
		else:  # min_black_list_matches <= 0:
			return len([cell for cell in column\
				if any(regex.fullmatch(cell)\
				for regex in Table.BLACKLISTED_COMPILED_REGEXES)])\
				>= len(column) + min_black_list_matches


	@classmethod
	def column_uniqueness(cls, column: List[str]) -> float:
		"""
		A helper function for get_identifying_column().
		Returns a score between 0.0 and 1.0 signifying how "unique" a column
		is (the score is 1.0 iff all cells in the column are unique;
		for constant columns the score tends towards 0.0).
		"""
		return len(set(column)) / len(column)


	@classmethod
	def column_is_unique(cls, column: List[str],\
		min_uniqueness: float = 0.5) -> bool:
		"""
		A helper function for get_identifying_column().
		Returns whether a column consists of at least
		min_uniqueness*100% unique values
		(i.e. only for min_uniqueness=1.0, the column is actually required
		 to consist of all truly unique values).
		For min_uniqueness=0.0, this function always returns True.
		"""
		return len(set(column)) >= min_uniqueness * len(column)


	def get_identifying_column(self, min_uniqueness: float = 0.5,\
		consider_non_textual_columns_as_last_resort: bool = False,\
		min_black_list_matches: int = 1) -> List[str]:
		"""
		Get (the extension of) the identifying column of this table.

		Ritze et al. call this identifying column the "entity label attribute"
		in their "Matching HTML Tables to DBpedia" paper (T2K Match).
		They use the following heuristic to identify it:
		(a) take the string attribute with the highest number of unique values
		(b) use the left-most attribute in case of a tie

		Quercini & Reynaud's paper
		"Entity Discovery and Annotation in Tables" however provides us with
		yet another valuable idea:
		ignoring cells whose content matches the regular expression for URLs,
		email addresses, phone numbers, geographic coordinates, etc.

		We combine these two ideas like this:
		(1) Is there exactly one unique column, ignoring other unique columns
		    that match one of the "blacklisted" regexes?
		    => if so, return that one.
		    (the fact that it matches none of the regexes implies
		     that it is a textual column)
		(2) Are there multiple unique columns, ignoring other unique columns
		    that match one of the "blacklisted" regexes?
		    => if so, return the left-most of these.
		(3) Return the column with the highest number of unique values,
		    ignoring other columns that match one of the "blacklisted" regexes,
		    but only if at least 50% of the values are unique.
		    In case of a tie, use the left-most attribute.
		(4) Only if consider_non_textual_columns_as_last_resort=True
		    (by default, it's set to False):
		    If no column reaches the 50% uniqueness value, also consider
		    columns that match one of the "blacklisted" regexes, but keeping
		    the requirement of 50% uniqueness.
		(5) No candidate for an identifying column was found.
		    Return the empty list [].

		Note that the 50% threshhold can be altered using the `min_uniqueness`
		parameter (float value between 0.0 and 1.0, default value = 0.5).
		With a `min_uniqueness` of 0.0 and
		`consider_non_textual_columns_as_last_resort` set to True,
		there will always be an identifying column found in step (3),
		or (4) when all columns match a "blacklisted" regex.

		The `min_black_list_matches` parameter can be used to adjust when a
		column is considers to "match" a blacklisted regex:
		* min_black_list_matches=1 (default):
		  A column is considered a match, even if just one cell in it
		  matches a blacklisted regex.
		* min_black_list_matches=n:
		  A column is considered a match, if at least n cells in it
		  match a blacklisted regex.
		* min_black_list_matches=0:
		  A column is considered a match, only if **all** cells in it
		  match a blacklisted regex.
		* min_black_list_matches=-1:
		  A column is considered a match, only if all cells, except one, in it
		  match a blacklisted regex.
		* min_black_list_matches=-n:
		  A column is considered a match, only if all cells, except n,
		  i.e. at least len(column)-n, in it match a blacklisted regex.
		"""

		blacklisted_columns: List[List[str]] = []
		non_blacklisted_columns: List[List[str]] = []
		for col in self.columns:
			if Table.column_is_blacklisted(col, min_black_list_matches):
				blacklisted_columns.append(col)
			else:
				non_blacklisted_columns.append(col)

		# (1) Is there exactly one unique column (ignoring blacklisted)?
		#     => if so, return that one.
		#
		# (2) Are there multiple unique columns (ignoring blacklisted)?
		#     => if so, return the left-most of these.
		unique_non_blacklisted_columns =\
			[col for col in non_blacklisted_columns\
			if Table.column_is_unique(col, min_uniqueness=1.0)]
		if len(unique_non_blacklisted_columns) > 0:
			return unique_non_blacklisted_columns[0]  # (return left-most col)

		# (3) Return the column with the highest number of unique values
		#     (ignoring blacklisted and at least min_uniqueness*100% unique,
		#      using the left-most in case of a tie):
		most_unique_columns: List[Tuple[List[str], float]] =\
			sorted([(col, Table.column_uniqueness(col))\
			for col in non_blacklisted_columns],\
			key=lambda tuple: tuple[1], reverse=True)
		if len(most_unique_columns) > 0\
			and most_unique_columns[0][1] >= min_uniqueness:
			return most_unique_columns[0][0]

		# (4) Only if consider_non_textual_columns_as_last_resort=True:
		#     Like step (3) but now (including) blacklisted columns:
		if consider_non_textual_columns_as_last_resort:
			most_unique_columns: List[Tuple[List[str], float]] =\
			sorted([(col, Table.column_uniqueness(col))\
			for col in blacklisted_columns],\
			key=lambda tuple: tuple[1], reverse=True)
			if len(most_unique_columns) > 0\
				and most_unique_columns[0][1] >= min_uniqueness:
				return most_unique_columns[0][0]

		# (5) No candidate for an identifying column was found.
		#     => Return the empty list [].
		return []


	def classifyGenerically(self,\
				 useSBERT:bool, useBing=False, useWebIsAdb=False,\
				 printProgressTo=sys.stdout, DEBUG=False)\
				 -> ClassificationResult:
		"""
		Just like classify(), only that the returned result is still generic,
		i.e. one can still determine which of the 3 approaches to use, whether
		to normalize them and how to weight them.

		This function is particularly useful for generating statistics,
		varying the parameters just listed.
		"""
		resultUsingTextualSurroundings: Dict[WikidataItem, float] = {}
		if self.surroundingText != "":
			print("[PROGRESS] Classifying using textual surroundings...",\
				file=printProgressTo)
			resultUsingTextualSurroundings =\
				self.classifyUsingTextualSurroundings()
			if DEBUG:
				print("[DEBUG] Result using textual surroundings: " +\
					f"{debug_dict_sorted(resultUsingTextualSurroundings)}")

		resultUsingAttrNames: Dict[WikidataItem, float] = {}
		if self.headerRow != []:
			print("[PROGRESS] Classifying using attribute names...",\
				file=printProgressTo)
			resultUsingAttrNames = self.classifyUsingAttrNames(\
				useSBERT=useSBERT)
			if DEBUG:
				print("[DEBUG] Result using attribute names: " +\
					f"{debug_dict_sorted(resultUsingAttrNames)}")


		resultUsingAttrExtensions: Dict[WikidataItem, float] = {}
		if self.columns != []:
			print("[PROGRESS] Classifying using attribute extensions...",\
				file=printProgressTo)
			resultUsingAttrExtensions = self.classifyUsingAttrExtensions(\
				useBing=useBing, useWebIsAdb=useWebIsAdb)
			if DEBUG:
				print("[DEBUG] Result using attribute extensions: " +\
					f"{debug_dict_sorted(resultUsingAttrExtensions)}")

		return ClassificationResult(\
			resUsingTextualSurroundings=resultUsingTextualSurroundings,\
			resUsingAttrNames=resultUsingAttrNames,\
			resUsingAttrExtensions=resultUsingAttrExtensions\
		)


	def classify(self,\
				 useSBERT:bool, useBing=False, useWebIsAdb=False,\
				 useTextualSurroundings=True, textualSurroundingsWeighting=1.0,\
				 useAttrNames=True, attrNamesWeighting=1.0,\
				 useAttrExtensions=True, attrExtensionsWeighting=1.0,\
				 normalizeApproaches=False,
				 printProgressTo=sys.stdout,\
				 DEBUG=False)\
				-> List[Tuple[float, WikidataItem]]:
		"""
		Classify this table semantically.
		Returns an ordered list of WikidataItems that might represent the entity
		type of the tuples of this table, each with a floating-point score
		indicating how well it fits the table.
		
    	Keyword arguments:
    	useTextualSurroundings -- whether to use approach #1 if possible
    	                          (default True)
    	textualSurroundingsWeighting -- how to weight approach #1
    	                                (default 0.0)
    	useAttrNames -- whether to use approach #2 if possible
    	                (default True)
    	attrNamesWeighting -- how to weight approach #1
                              (default 0.0)
    	useAttrExtensions -- whether to use approach #3 if possible
    	                     (default True)
    	attrExtensionsWeighting -- how to weight approach #1
    	                           (default 0.0)
    	normalizeApproaches -- whether to normalize the result of each of the 3
    	                       approaches into the [0,1] range
    	                       (default False)

		More about `normalizeApproaches`:

    	Whether to normalize the rankings of classifyUsingTextualSurroundings(),
	    classifyUsingAttrNames(), classifyUsingAttrExtensions() into the same
	    range, i.e. the interval [0,1].
	 
	    When normalizeApproaches == False, the final ranking of classify()
	    is computed as:
	    classify() =
	        TEXTUAL_SURROUNDINGS_WEIGHTING * classifyUsingTextualSurroundings()
	      + ATTR_NAMES_WEIGHTING * classifyUsingAttrNames()
	      + ATTR_EXTENSIONS_WEIGHTING * classifyUsingAttrExtensions()
	 
	    When normalizeApproaches == True, the final ranking of classify()
	    is computed as:
	    classify() =
	        TXT_SURR_WGHT * normalize(classifyUsingTextualSurroundings())
	      + ATTR_NAMES_WEIGHTING * normalize(classifyUsingAttrNames())
	      + ATTR_EXTENSIONS_WEIGHTING * normalize(classifyUsingAttrExtensions())
	    => Advantage:
	       The weightings truly reflect the contribution of each approach.
	    => Disadvantage:
	       The information that one approach might be more confident than
	       another approach is lost.

	    `printProgressTo` may be set to `sys.stdout`, `sys.stderr` or
	    `open(os.devnull,"w")` for example.
    	"""

		resultUsingTextualSurroundings: Dict[WikidataItem, float] = {}
		if useTextualSurroundings and self.surroundingText != "":
			print("[PROGRESS] Classifying using textual surroundings...",\
				file=printProgressTo)
			resultUsingTextualSurroundings =\
				self.classifyUsingTextualSurroundings()
			if DEBUG:
				print("[DEBUG] Result using textual surroundings: " +\
					f"{debug_dict_sorted(resultUsingTextualSurroundings)}")

		resultUsingAttrNames: Dict[WikidataItem, float] = {}
		if useAttrNames and self.headerRow != []:
			print("[PROGRESS] Classifying using attribute names...",\
				file=printProgressTo)
			resultUsingAttrNames = self.classifyUsingAttrNames(\
				useSBERT=useSBERT)
			if DEBUG:
				print("[DEBUG] Result using attribute names: " +\
					f"{debug_dict_sorted(resultUsingAttrNames)}")


		resultUsingAttrExtensions: Dict[WikidataItem, float] = {}
		if useAttrExtensions and self.columns != []:
			print("[PROGRESS] Classifying using attribute extensions...",\
				file=printProgressTo)
			resultUsingAttrExtensions = self.classifyUsingAttrExtensions(\
				useBing=useBing, useWebIsAdb=useWebIsAdb)
			if DEBUG:
				print("[DEBUG] Result using attribute extensions: " +\
					f"{debug_dict_sorted(resultUsingAttrExtensions)}")

		return ClassificationResult(\
			resUsingTextualSurroundings=resultUsingTextualSurroundings,\
			resUsingAttrNames=resultUsingAttrNames,\
			resUsingAttrExtensions=resultUsingAttrExtensions\
		).classify(\
			useTextualSurroundings=useTextualSurroundings,\
			textualSurroundingsWeighting=textualSurroundingsWeighting,\
			useAttrNames=useAttrNames,\
			attrNamesWeighting=attrNamesWeighting,\
			useAttrExtensions=useAttrExtensions,\
			attrExtensionsWeighting=attrExtensionsWeighting,\
			normalizeApproaches=normalizeApproaches,
			DEBUG=DEBUG\
		)

	def classifyUsingTextualSurroundings(self) -> Dict[WikidataItem, float]:
		"""
		Classify this table using the "Using Textual Surroundings" approach,
		implemented in filter_nouns_with_heuristics.py.
		Higher scores mean better matches.
		"""

		return filter_nouns_with_heuristics_as_dict(\
			input_text=self.surroundingText,\
			VERBOSE=False,\
			ALLOW_EQUAL_SCORES=True)
		# (already returns a Dict[WikidataItem, float])

	def classifyUsingAttrNames(self, useSBERT: bool)\
		-> Dict[WikidataItem, float]:
		"""
		Classify this table using the "Using Attribute Names" approach,
		implemented in attr_names_to_ontology_class.py.
		Higher scores mean better matches.

		When useSBERT==False, the Jaccard index is used for computing
		word similarity instead.
		"""

		# First, clean the header row / attribute names:
		# (a) remove empty column captions ("")
		# (b) remove column captions shorter than 3 characters when using
		#     Jaccard-trigrams instead of SBERT
		cleaned_headerRow: List[str] =\
			[header for header in self.headerRow\
			 if header.strip() != "" and (len(header) >= 3 or useSBERT)]

		# attr_names_to_ontology_class() returns a Dict[str, float]
		# (with the strings being names of DBpedia classes)
		# but we need a Dict[WikidataItem, float]:
		return { WikidataItem(\
			get_dbpedia_classes_mapped_to_wikidata()[dbpediaClass]) : score\
			for dbpediaClass, score\
			in attr_names_to_ontology_class(\
				inputAttrNames=cleaned_headerRow,
				USE_BETTER_SUM_FORMULA=True,
				USE_SBERT_INSTEAD_OF_JACCARD=useSBERT,
				VERBOSE=False
			).items()}

	def classifyUsingAttrExtensions(self, useBing=False, useWebIsAdb=False)\
		 -> Dict[WikidataItem, float]:
		"""
		Classify this table using the "Using Attribute Extensions" approach,
		implemented in attr_extension_to_ontology_class.py.
		Higher scores mean better matches.

		The `useBing` and `useWebIsAdb` parameters
		shall not both be set to True!
		"""

		if useBing:
			return\
				attr_extension_to_ontology_class_web_search_list_onto_as_dict(\
				cell_labels=self.get_identifying_column())
			# (already returns a Dict[WikidataItem, float])
		elif useWebIsAdb:
			exit("[ERROR] Using the WebIsA DB is not yet implemented!")
			return None  # (possible ToDo: not even coded yet...)
		else:
			return\
				attr_extension_to_ontology_class(\
				cell_labels=self.get_identifying_column())
			# (returns a Dict[WikidataItem, int])


	def has_co_occurring_keywords(self, keywords: List[str], requireAll=False,\
		lookInSurroundingText=True, lookInsideTable=True) -> bool: # ToDo: USE !!!
		"""
		Use narrative knowledge:

		Check whether this table has at least one of the keywords specified
		in `keywords` occuring in its surrounding text or inside its content.

		Set `requireAll` to True to require all keywords specified in
		`keywords` to occur in the surrounding text or table content
		(note that setting `requireAll` to True makes this function slower!).

		Set `lookInSurroundingText` or `lookInsideTable` to False to **only**
		look into the surrounding text or table content.

		Example:

		t = Table("Hello world, this is the surrounding text!",
		["header_foo", "header_bar", "header_baz"],
		[["1", "2", "3"], ["one", "two", "three"], ["un", "dos", "tres"]])

		>>> t.has_co_occurring_keywords(["world"])
		True
		>>> t.has_co_occurring_keywords(["foo"])
		True
		>>> t.has_co_occurring_keywords(["tres"])
		True
		>>> 
		>>> 
		>>> t.has_co_occurring_keywords(["4", "3", "0"])
		True
		>>> t.has_co_occurring_keywords(["4", "3", "0"], requireAll=True)
		False
		>>> t.has_co_occurring_keywords(["2", "3", "1"], requireAll=True)
		True
		>>> t.has_co_occurring_keywords(["2", "3", "1", "1"], requireAll=True)
		True
		>>> 
		>>> 
		>>> t.has_co_occurring_keywords(["2", "3", "1", "1"], requireAll=True,\
		... lookInsideTable=False)
		False
		>>> t.has_co_occurring_keywords(["world"], lookInSurroundingText=False)
		False
		>>> t.has_co_occurring_keywords(["foo"], lookInsideTable=False)
		False
		>>> t.has_co_occurring_keywords(["tres"], lookInsideTable=False)
		False
		"""

		# The set of all keywords in `keywords` that were found in the
		#   surrounding text of this table and/or its content:
		keywords_found_set: Set[str] = set()  # a subset of `keywords`

		# (1) Look in self.surroundingText:
		if lookInSurroundingText:
			for keyword in keywords:
				if keyword in self.surroundingText:
					keywords_found_set.add(keyword)
					if not requireAll:
						# If finding one keyword is enough, we can immediately
						#  terminate here:
						return True

		# (2) Look inside this table, i.e. self.headerRow and self.columns:
		if lookInsideTable:
			for keyword in keywords:
				if any(keyword in header\
						for header in self.headerRow) or\
					any(keyword in cell\
						for column in self.columns for cell in column):
					keywords_found_set.add(keyword)
					if not requireAll:
						# If finding one keyword is enough, we can immediately
						#  terminate here:
						return True
		
		if requireAll:  # All keywords in `keywords` have to be found:
			return len(keywords_found_set) == len(set(keywords))
		else:  # At least one keyword in `keywords` has to be found:
			return len(keywords_found_set) > 0

	def fulfills_attribute_condition(self, attribute_cond: str,\
		useSBERT: bool = True, strictness: float = 1.0,\
		DEBUG: bool = False) -> bool:
		"""
		Use narrative knowledge:

		Check whether this Table
		(a) has the attribute named in `attribute_cond`
		    (name doesn't have to be exactly equal, SBERT/Jaccard is used for
		    attribute name similarity here again), and
		(b) all (or less, when strictness < 1.0) values in the corresponding
		    column of this Table fulfill the specified `attribute_cond`.
		Returns either True or False.

		(a) and (b) are used in conjunction, i.e. when a column fulfills the
		requirement in (b) it is considered a more likely match for (a) !

		Example:
		Only look for tables containing heavy vehicles or super sports cars:
		table.fulfills_attribute_condition("horsepower >= 500", useSBERT=True)

		Arguments:
		attribute_cond -- a string, boolean condition in Python syntax:
                          "[attribute name] [<=,>=,<,>,==,!=,in,not in]
                           [value, value range or value list]", e.g.
                          "horsepower >= 500" or
                          "year in range(1980,2000)" or
                          "firstName not in ['Alex', 'Alexander']"
        useSBERT -- whether to use SBERT to match the attribute name given
                    in the `attribute_cond` parameter to this tables headers;
                    otherwise Jaccard trigram similarity is used
		strictness -- a value between 0.0 (least strict)
		              and 1.0 (most strict, default); the fraction of cells
		              in the column that have to fulfill the condition
		              (by default, every cell has to fulfill the condition)
		"""

		# (0) When this table has no header, no meaningful matching is possible:
		if self.headerRow == []:
			return False

		# (0) The similarity function used:
		def similarity(attrName1: str, attrName2: str) -> float:
			return attr_names_to_ontology_class.similarity(\
				attrName1=attrName1, attrName2=attrName2, USE_SBERT=useSBERT)

		# (1) Extract the attribute name from the given attribute condition:
		attribute_name: str = attribute_cond.split()[0]

		# (2) Turn the condition into a Python lambda (for individual cells):
		attribute_cond_lambda =\
			eval(f"lambda {attribute_name}: {attribute_cond}")
		def attribute_cond_lambda_type_safe(cell_value):
			try:
				# First, try to insert the cell value as the string that it is:
				return attribute_cond_lambda(cell_value)
			except:
				try:
					# Second, try to cast the cell value to a float:
					return attribute_cond_lambda(float(cell_value))
				except:
					# Third, try to turn the cell value to a bool,
					#   else return False:
					try:
						if cell_value.lower() in ["true", "y", "yes"]:
							return attribute_cond_lambda(True)
						elif cell_value.lower() in ["false", "n", "no"]:
							return attribute_cond_lambda(False)
						else:
							return False
					except:
						return False  # (give up)


		# (3) Turn the condition into a Python lambda (for entire columns):
		attribute_cond_lambda_column =\
			lambda column: len([cell for cell in column\
				if attribute_cond_lambda_type_safe(cell)])\
				>= strictness * len(column)

		# (4) Do the first (a) matching:
		a_matching: Dict[str, float] = {column_name:\
			similarity(column_name, attribute_name)\
			for column_name in self.headerRow}
		if DEBUG: print(f"[DEBUG] first (a) matching: {a_matching}")

		# (5) Do the (b) matching:
		b_matching: Dict[str, bool] = {self.headerRow[x]:\
			attribute_cond_lambda_column(self.columns[x])\
			for x in range(0, len(self.columns))}
		if DEBUG: print(f"[DEBUG] (b) matching: {b_matching}")

		# (6) When there are no (b) matches at all, return False:
		if not any(b_matching.values()):
			if DEBUG: print("[DEBUG] no (b) matches at all => ret False")
			return False

		# (7) When the best (a) match is also a (b) match, return True:
		best_a_match: str = max(a_matching, key=a_matching.get)
		if DEBUG: print(f"[DEBUG] best (a) match = '{best_a_match}'")
		if b_matching[best_a_match]:
			if DEBUG: print("[DEBUG] best (a) match is (b) match => ret True")
			return True

		# (8) Increase the (a) scores (by 0.1) where there's a (b) match.
		#     If the best (a) match *then* is also a (b) match, return True,
		#     otherwise False:
		a_matching = {k : v + (0.1 if b_matching[k] else 0.0)\
			for k, v in a_matching.items()}
		if DEBUG: print(f"[DEBUG] second (a) matching: {a_matching}")
		best_a_match: str = max(a_matching, key=a_matching.get)
		if DEBUG: print(f"[DEBUG] best (a) match = '{best_a_match}'")
		if b_matching[best_a_match]:
			if DEBUG: print("[DEBUG] best (a) match is (b) match => ret True")
			return True
		else:
			if DEBUG: print("[DEBUG] best (a) match no (b) match => ret False")
			return False


	@classmethod
	def identifyCSVdialect(cls, csv_str: str) -> Dialect:
		return csv.Sniffer().sniff(csv_str)

	@classmethod
	def parseCSV(cls, csv_str: str, dialect: Dialect = None) -> Table:
		columns: List[List[str]] = []

		# Documentation of
		#   csv.reader(csvfile, dialect='excel', **fmtparams):
		# "csvfile can be any object which supports the iterator protocol
		#  and returns a string each time its __next__() method is called —
		#  file objects and list objects are both suitable. If csvfile is a
		#  file object, it should be opened with newline=''."

		csv_reader = csv.reader(csv_str.splitlines(keepends=True), dialect\
			if dialect is not None else Table.identifyCSVdialect(csv_str))
		for row in csv_reader:
			no_of_cols: int = len(row)  # (number of columns)
			if columns == []:
				columns = [[] for i in range(0, no_of_cols)]  # (initialize)
			for col_index in range(0, no_of_cols):
				columns[col_index].append(row[col_index])

		result = Table(surroundingText="", headerRow=[],\
			columns=columns)  # surroundingText is not applicable to CSV files
		result.parse_header_row()  # set headerRow!=[] if possible
		return result

	@classmethod
	def parseXLSX(cls, xlsxPath: str) -> Table:
		# Use openpyxl to parse Microsoft Excel files,
		# cf. https://www.geeksforgeeks.org/
		#   working-with-excel-spreadsheets-in-python/

		import openpyxl
		wb_obj = openpyxl.load_workbook(xlsxPath)
		sheet_obj = wb_obj.active

		# # Example from www.geeksforgeeks.org :
		# cell_obj = sheet_obj.cell(row = 1, column = 1)
		# print(cell_obj.value)

		# # Works but does not consider a possible header row:
		# columns = [list(map(lambda cell_obj: str(cell_obj.value), column))\
		# 	for column in sheet_obj.iter_cols()]

		rows = [list(map(lambda cell_obj: str(cell_obj.value), row))\
			for row in sheet_obj.iter_rows()]

		#If Table.parseCSV() took an iterator instead, it would look like this:
		#return Table.parseCSV(csv_lines=(",".join(row) + "\n" for row in rows))
		csv_str = "\n".join([",".join(row) for row in rows])
		return Table.parseCSV(csv_str)

	@classmethod
	def parseJSON(cls, json_str: str, onlyRelational=False,\
		onlyWithHeader=True, useAdditionalHeuristics=False,\
		DEBUG=False) -> Optional[Table]: # ToDo: set onlyRelational according to cmd line flag!!
		"""
		Parses a .JSON file of the format that's used in the
		WDC Web Table Corpus (http://webdatacommons.org/webtables/2015/
		downloadInstructions.html):

		{"relation":[[..., ...],[..., ...]],
		 "pageTitle":"...",
		 "title":"...",
		 "url":"https://...",
		 "hasHeader":true,
		 "headerPosition":"MIXED",
		 "tableType":"RELATION",
		 "tableNum":6,
		 "s3Link":"common-crawl/crawl-data/...",
		 "recordEndOffset":867273895,
		 "recordOffset":867263717,
		 "tableOrientation":"HORIZONTAL",
		 "textBeforeTable":"...",
		 "textAfterTable":"...",
		 "hasKeyColumn":true,
		 "keyColumnIndex":0,
		 "headerRowIndex":0}

		This method returns None for invalid JSON inputs
		(i.e. not of the above format), for non-relational tables
		(if onlyRelational=True) and for tables without a header
		(if onlyWithHeader=True)!
		"""
		try:
			j = json.loads(json_str)
			
			if onlyRelational and j["tableType"] != "RELATION":
				# e.g. "tableType":"ENTITY"
				return None  # skip non-relational tables

			surroundingText: str =\
				j["textBeforeTable"] + " " + j["textAfterTable"]
			headerRow: List[str] = None
			columns: List[List[str]] = None

			# For the header information, rely on the information given
			#   (i.e. do no own heuristics):
			if j["hasHeader"]:
				if j["headerPosition"] == "FIRST_ROW":  # (most common case)
					headerRow = [col[0] for col in j["relation"]]
					columns = [col[1:] for col in j["relation"]] # strip header
				elif j["headerPosition"] == "FIRST_COLUMN":
					# Tranpose the table because for us, the header is always
					#   in the first row:
					headerRow = j["relation"][0]  # 1st column = header
					columns = transpose(j["relation"][1:]).tolist()
				elif j["headerPosition"] == "MIXED":
					return None  # skipping such unsure/non-relational? tables
				else:
					print("[PARSE ERROR] Unknown value for 'headerPosition'" +\
						f""" in JSON: '{j["headerPosition"]}'""",\
						file=sys.stderr)
					return None
			else:  # "hasHeader":false => "headerPosition":"NONE"
				if onlyWithHeader and not useAdditionalHeuristics:
					# Corpus says that this table has no header, we shall
					#   not use additional heuristcs and we shall only
					#   return tables with header => return None:
					return None
				elif not onlyWithHeader and not useAdditionalHeuristics:
					# Corpus says that this table has no header, we **are**
					#   allowed to return tables without a header but we are
					#   **not** allowed to use additional heuristics
					#   => return the table without header
					#      as it is in the corpus:
					headerRow = []
					# We still have to find out which way to orient the columns
					#   though:
					columns = None  # ToDo(!)
				elif useAdditionalHeuristics:
					# We **are** allowed to use additional heuristics:
					# Looking at some examples shows that when
					#   "hasHeader":false, there **is** usually a header
					#   in the first column...
					return None  # ToDo(!)

			return Table(surroundingText=surroundingText, headerRow=headerRow,\
				columns=columns)
		except BaseException as e:
			# (most likely a KeyError when JSON is in invalid format)
			if DEBUG:
				print(f"[DEBUG] Error parsing JSON: {e}", file=sys.stderr)
			return None

	@classmethod
	def parseTAR(cls, tarPath: str, file_extensions: FileExtensions,\
		csv_dialect: Dialect = None, onlyRelationalJSON=False, DEBUG=False)\
		-> Iterator[Table]:
		# cf. https://docs.python.org/3/library/tarfile.html
		tar = tarfile.open(tarPath)
		for tarinfo in tar:
			# Skip directories and hidden files in the TAR archive:
			if tarinfo.isfile() and tarinfo.name[:1] != ".":
				table: Table = None
				if os.path.splitext(tarinfo.name)[1]\
					in file_extensions.JSON_extensions:
					table = Table.parseJSON(\
						json_str=tar.extractfile(tarinfo).read(),\
						onlyRelational=onlyRelationalJSON,\
						DEBUG=DEBUG)
				elif os.path.splitext(tarinfo.name)[1]\
					in file_extensions.CSV_extensions:
					table = Table.parseCSV(\
						csv_str=tar.extractfile(tarinfo).read(),\
						dialect=csv_dialect)
				elif os.path.splitext(tarinfo.name)[1]\
					in file_extensions.XLSX_extensions:
					# Because parseXLSX() expects a path, we have to
					#   temporarily extract the .xlsx file to the file system:
					with tempfile.NamedTemporaryFile() as temp_path:
						# Extract the .xlsx file to a temporary location:
						tar.extract(tarinfo, path=temp_path)
						# Parse that temporary .xlsx file:
						table = Table.parseXLSX(xlsxPath=temp_path)
				if table is not None: table.file_name = tarinfo.name
				yield table
				# All other file types in the TAR archive are ignored.


	@classmethod
	def parseFolder(cls, folderPath: str, file_extensions: FileExtensions,\
		recursive=True, csv_dialect: Dialect = None, onlyRelationalJSON=False,\
		DEBUG=False) -> Iterator[Table]:
		# Sorting is important such that the order is predictable to the user:
		for folder_item in sorted(os.listdir(folderPath)):
			folder_item = os.path.join(folderPath, folder_item)
			if os.path.isfile(folder_item):
				file_name, file_extension = os.path.splitext(folder_item)
				if file_extension in file_extensions.CSV_extensions:
					with open(folder_item, 'r', newline='') as csv_file:
						# (newline='' is required by the CSV library)
						table = Table.parseCSV(csv_str=csv_file.read(),\
							dialect=csv_dialect)
						if table is not None: table.file_name = folder_item
						yield table
				elif file_extension in file_extensions.XLSX_extensions:
					table = Table.parseXLSX(xlsxPath=folder_item)
					if table is not None: table.file_name = folder_item
					yield table
				elif file_extension in file_extensions.JSON_extensions:
					with open(folder_item, 'r') as json_file:
						table = Table.parseJSON(json_str=json_file.read(),\
							onlyRelational=onlyRelationalJSON,\
							DEBUG=DEBUG)
						if table is not None: table.file_name = folder_item
						yield table
				elif file_extension in file_extensions.TAR_extensions:
					for table in Table.parseTAR(tarPath=folder_item,\
						file_extensions=file_extensions,\
						csv_dialect=csv_dialect,\
						onlyRelationalJSON=onlyRelationalJSON, DEBUG=DEBUG):
						yield table
				# Skip files with any other filetype/extension.
			elif recursive:
				for table in Table.parseFolder(folderPath=folder_item,\
					recursive=True, csv_dialect=csv_dialect,\
					onlyRelationalJSON=onlyRelationalJSON, DEBUG=DEBUG):
					yield table
			# When the item is a directory and recursive=False, skip it.

	@classmethod
	def parseCorpus(cls, corpusPath: str, file_extensions: FileExtensions,\
		recursive=True, yieldNones=False, csv_dialect: Dialect = None,\
		onlyRelationalJSON=False, min_table_size: int = 1,\
		DEBUG=False) -> Iterator[Table]:
		table_filter: Callable[Table, bool] =\
			lambda table: (yieldNones or table is not None)\
				and (table is None or table.min_dimension() >= min_table_size)

		if os.path.isfile(corpusPath):
			file_name, file_extension = os.path.splitext(corpusPath)
			if file_extension in file_extensions.TAR_extensions:
				return filter(table_filter,\
					Table.parseTAR(tarPath=corpusPath,\
						file_extensions=file_extensions,\
						csv_dialect=csv_dialect,\
						onlyRelationalJSON=onlyRelationalJSON, DEBUG=DEBUG))
			else:
				sys.exit("Error: --corpus supplied either has to be a "+\
					" directory or a .TAR file.")
		else:
			return filter(table_filter,\
				Table.parseFolder(folderPath=corpusPath,\
					file_extensions=file_extensions, recursive=recursive,\
					csv_dialect=csv_dialect,\
					onlyRelationalJSON=onlyRelationalJSON, DEBUG=DEBUG))

	@classmethod  # ToDo: USE !!!
	def create2DStatisticalTable(cls,\
		_lambda: Callable[[float, float], str],\
		x_var_name: str = "x",\
		y_var_name: str = "y",\
		x_range: List[float] =\
			[0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0],\
		y_range: List[float] =\
			[0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0],\
		) -> Table:
		"""
		This function is for creating statistical tables to then print them out
		using the pretty_print() function.
		This has nothing to do with corpus tables directly.

		# Example:

		t = Table.create2DStatisticalTable(lambda x, y: "{:6.3f}".format(x-y))

		>>> print(t.pretty_print())
		      | x=0.0  | x=0.1  | x=0.2  | x=0.3  | x=0.4  | x=0.5  | ... 
		------------------------------------------------------------------
		y=0.0 |  0.000 |  0.100 |  0.200 |  0.300 |  0.400 |  0.500 | ... 
		y=0.1 | -0.100 |  0.000 |  0.100 |  0.200 |  0.300 |  0.400 | ... 
		y=0.2 | -0.200 | -0.100 |  0.000 |  0.100 |  0.200 |  0.300 | ... 
		y=0.3 | -0.300 | -0.200 | -0.100 |  0.000 |  0.100 |  0.200 | ... 
		y=0.4 | -0.400 | -0.300 | -0.200 | -0.100 |  0.000 |  0.100 | ... 
		y=0.5 | -0.500 | -0.400 | -0.300 | -0.200 | -0.100 |  0.000 | ... 
		...   | ...    | ...    | ...    | ...    | ...    | ...    | ...   

		# The same example with custom variable names:

		t = Table.create2DStatisticalTable(lambda x, y: "{:6.3f}".format(x-y),
		x_var_name="w1", y_var_name="w2")

		>>> print(t.pretty_print())
		       | w1=0.0 | w1=0.1 | w1=0.2 | w1=0.3 | w1=0.4 | w1=0.5 | ...
		-------------------------------------------------------------------
		w2=0.0 |  0.000 |  0.100 |  0.200 |  0.300 |  0.400 |  0.500 | ...
		w2=0.1 | -0.100 |  0.000 |  0.100 |  0.200 |  0.300 |  0.400 | ...
		w2=0.2 | -0.200 | -0.100 |  0.000 |  0.100 |  0.200 |  0.300 | ...
		w2=0.3 | -0.300 | -0.200 | -0.100 |  0.000 |  0.100 |  0.200 | ...
		w2=0.4 | -0.400 | -0.300 | -0.200 | -0.100 |  0.000 |  0.100 | ...
		w2=0.5 | -0.500 | -0.400 | -0.300 | -0.200 | -0.100 |  0.000 | ...
		...    | ...    | ...    | ...    | ...    | ...    | ...    | ...
		"""

		return Table(\
			surroundingText="",\
			headerRow = [""] + [f"{x_var_name}={x}" for x in x_range],\
			columns=[[f"{y_var_name}={y}" for y in y_range]] +\
				[[str(_lambda(x,y)) for y in y_range] for x in x_range]
			)

	@classmethod  # ToDo: USE !!!!!!
	def create3DStatisticalTable(cls,\
		_lambda: Callable[[float, float, float], str],\
		x_var_name: str = "x",\
		left_y_var_name: str = "y1",\
		right_y_var_name: str = "y2",\
		x_range: List[float] =\
			[0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0],\
		left_y_range: List[float] =\
			[0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0],\
		right_y_range: List[float] =\
			[0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]\
		) -> Table:
		"""
		This function is for creating statistical tables to then print them out
		using the pretty_print() function.
		This has nothing to do with corpus tables directly.

		Example:

		t = Table.create3DStatisticalTable(lambda w1, k, w2:
			"{:6.3f}".format(k*(w1-w2)),
			x_var_name="w1",
			left_y_var_name="k",
			right_y_var_name="w2",
			x_range=[0.0, 0.1, 0.2, 0.3, 0.4, 0.5],
			left_y_range = [1,2,3],
			right_y_range = [0.0, 0.1, 0.2, 0.3])
		
		>>> print(t.pretty_print(3*4))
		    |        | w1=0.0 | w1=0.1 | w1=0.2 | w1=0.3 | w1=0.4 | w1=0.5
		------------------------------------------------------------------
		k=1 | w2=0.0 |  0.000 |  0.100 |  0.200 |  0.300 |  0.400 |  0.500
		k=1 | w2=0.1 | -0.100 |  0.000 |  0.100 |  0.200 |  0.300 |  0.400
		k=1 | w2=0.2 | -0.200 | -0.100 |  0.000 |  0.100 |  0.200 |  0.300
		k=1 | w2=0.3 | -0.300 | -0.200 | -0.100 |  0.000 |  0.100 |  0.200
		k=2 | w2=0.0 |  0.000 |  0.200 |  0.400 |  0.600 |  0.800 |  1.000
		k=2 | w2=0.1 | -0.200 |  0.000 |  0.200 |  0.400 |  0.600 |  0.800
		k=2 | w2=0.2 | -0.400 | -0.200 |  0.000 |  0.200 |  0.400 |  0.600
		k=2 | w2=0.3 | -0.600 | -0.400 | -0.200 |  0.000 |  0.200 |  0.400
		k=3 | w2=0.0 |  0.000 |  0.300 |  0.600 |  0.900 |  1.200 |  1.500
		k=3 | w2=0.1 | -0.300 |  0.000 |  0.300 |  0.600 |  0.900 |  1.200
		k=3 | w2=0.2 | -0.600 | -0.300 |  0.000 |  0.300 |  0.600 |  0.900
		k=3 | w2=0.3 | -0.900 | -0.600 | -0.300 |  0.000 |  0.300 |  0.600
		"""

		return Table(\
			surroundingText="",\
			headerRow = ["", ""] + [f"{x_var_name}={x}" for x in x_range],\
			columns=[[f"{left_y_var_name}={y}" for y in left_y_range\
						for i in range(len(right_y_range))]] +\
				[[f"{right_y_var_name}={y}"\
						for y in right_y_range] * len(left_y_range)] +\
				[[str(_lambda(x, left_y, right_y))\
					for left_y in left_y_range\
					for right_y in right_y_range]\
					for x in x_range]
			)

